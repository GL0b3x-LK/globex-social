"""Phase 0 — Asset Import & Inventory Gap Closure.

One-shot, idempotent, re-runnable. Pulls every usable source asset into the
project tree in the right format:

  * Converts the 3 vector EPS logos -> transparent PNG via Ghostscript's
    ``pngalpha`` device (renders the PostScript, not the embedded low-res TIFF
    preview, and yields a real alpha channel). Width-normalised to 2160px.
    A reversed "white" variant (navy -> white, cyan kept) is derived for dark
    template backgrounds.
  * Parses the Globex Excel (data_only) into employees / trade_shows / holidays
    JSON. Persists hire_date ONLY for employees -- never DOB, age, or passwords.
  * Stages sample photos, the contract PDF, and a client-chat reference stub.
  * Writes docs/missing_assets.md -- the consolidated gap checklist for Karen.

Run:  .venv\\Scripts\\python.exe scripts\\import_assets.py

Requires ImageMagick + Ghostscript (installed via Scoop). Tool discovery below
prepends the Scoop locations to PATH and sets the MAGICK_* env vars so the
portable ImageMagick build can find its coder modules + config.
"""

from __future__ import annotations

import json
import logging
import os
import re
import shutil
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = Path(r"C:\Users\abdur\Downloads\Globex\Globex")
EXCEL_PATH = Path(r"C:\Users\abdur\Downloads\Date of Hire Globex, Birthdate and PWs - Copy.xlsx")

DATA_DIR = PROJECT_ROOT / "app" / "data"
LOGO_DIR = PROJECT_ROOT / "app" / "templates" / "assets" / "logos"
PHOTO_FIXTURE_DIR = PROJECT_ROOT / "tests" / "fixtures" / "photos"
DOCS_DIR = PROJECT_ROOT / "docs"
DOCS_REFERENCE_DIR = DOCS_DIR / "reference"

BRAND_NAVY = "#002D72"  # Pantone 288C

# EPS filename -> output PNG stem
LOGO_MAP = {
    "Globex_G-Man_Pantone-288+2985 (2).eps": "globex-gman-full",
    "Globex_logo_side-lockup_Pantone-288+2985.eps": "globex-lockup-side",
    "Globex_logo_top-Globex-only_Pantone-288.eps": "globex-wordmark-navy",
}

TARGET_WIDTH = 2160  # final logo width in px (per plan / retina)
RENDER_WIDTH = 2592  # oversample width before downscale (~1.2x)

log = logging.getLogger("import_assets")


# --------------------------------------------------------------------------- #
# Tooling
# --------------------------------------------------------------------------- #
def setup_tools() -> tuple[str, str]:
    """Locate gswin64c + magick, wiring Scoop paths and MAGICK_* env vars.

    The tool host inherits a stale environment from session start, so Scoop's
    registry PATH edits aren't visible. Prepend the known Scoop locations and
    point ImageMagick at its coder modules / config explicitly.
    """
    scoop = Path.home() / "scoop"
    im_dir = scoop / "apps" / "imagemagick" / "current"
    extra = [
        str(scoop / "shims"),
        str(im_dir),
        str(scoop / "apps" / "ghostscript" / "current" / "bin"),
    ]
    os.environ["PATH"] = os.pathsep.join(extra) + os.pathsep + os.environ.get("PATH", "")
    if im_dir.exists():
        os.environ.setdefault("MAGICK_HOME", str(im_dir))
        os.environ["MAGICK_CONFIGURE_PATH"] = str(im_dir)
        os.environ["MAGICK_CODER_MODULE_PATH"] = str(im_dir / "modules" / "coders")

    gs = shutil.which("gswin64c") or shutil.which("gs")
    magick = shutil.which("magick")
    if not gs:
        sys.exit("FATAL: Ghostscript (gswin64c) not found. Install via: scoop install ghostscript")
    if not magick:
        sys.exit("FATAL: ImageMagick (magick) not found. Install via: scoop install imagemagick")
    return gs, magick


def _run(cmd: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, check=True, capture_output=True, text=True)


# --------------------------------------------------------------------------- #
# Logo conversion
# --------------------------------------------------------------------------- #
def read_bbox(src: Path) -> tuple[float, float]:
    """Return (width_pts, height_pts) from the EPS BoundingBox header."""
    head = src.read_bytes()[:8192].decode("latin-1", "ignore")
    m = re.search(r"%%HiResBoundingBox:\s*([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)", head)
    if not m:
        m = re.search(r"%%BoundingBox:\s*(-?\d+)\s+(-?\d+)\s+(-?\d+)\s+(-?\d+)", head)
    if not m:
        raise ValueError(f"No BoundingBox found in {src.name}")
    x0, y0, x1, y1 = map(float, m.groups())
    return (x1 - x0, y1 - y0)


def identify(magick: str, png: Path) -> str:
    r = _run([magick, "identify", "-format", "%wx%h alpha=%A", str(png)])
    return r.stdout.strip()


def convert_logo(gs: str, magick: str, src: Path, stem: str) -> dict:
    """EPS -> transparent full-colour PNG + reversed white variant."""
    bbox_w, _ = read_bbox(src)
    dpi = max(150, min(int(round(RENDER_WIDTH * 72 / bbox_w)), 3000))

    out_full = LOGO_DIR / f"{stem}.png"
    out_white = LOGO_DIR / f"{stem}-white.png"
    tmp = LOGO_DIR / f"{stem}.tmp.png"

    # 1. Render the vector PostScript to a transparent PNG (ignores TIFF preview).
    _run(
        [
            gs,
            "-q",
            "-dSAFER",
            "-dBATCH",
            "-dNOPAUSE",
            "-dEPSCrop",
            "-sDEVICE=pngalpha",
            f"-r{dpi}",
            f"-sOutputFile={tmp}",
            str(src),
        ]
    )
    # 2. Normalise to target width, strip metadata.
    _run([magick, str(tmp), "-resize", f"{TARGET_WIDTH}x", "-strip", str(out_full)])
    # 3. Reversed variant for dark backgrounds: recolour navy -> white, keep cyan.
    _run(
        [
            magick,
            str(out_full),
            "-fuzz",
            "20%",
            "-fill",
            "white",
            "-opaque",
            BRAND_NAVY,
            "-strip",
            str(out_white),
        ]
    )
    tmp.unlink(missing_ok=True)

    return {"stem": stem, "dpi": dpi, "full": identify(magick, out_full)}


# --------------------------------------------------------------------------- #
# Excel parsing helpers
# --------------------------------------------------------------------------- #
def _iso(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def parse_employees(wb) -> list[dict]:
    """Employee Info tab. hire_date ONLY -- DOB / age / passwords are dropped."""
    ws = wb["Employee Info"]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[1]
        if not name or not str(name).strip():
            continue
        out.append(
            {
                "name": _clean(name),
                "title": _clean(row[5]) if row[5] else None,
                "hire_date": _iso(row[3]),  # col D — Date of Hire
                "department": None,
                "active": True,
            }
        )
    return out


def _parse_event_dates(raw) -> tuple[str | None, str | None, bool]:
    """Parse 'M/D/YY to M/D/YY' / 'TBC' / blank -> (start, end, needs_confirmation)."""
    s = _clean(raw) if raw else ""
    if not s:
        return None, None, False
    if s.upper() == "TBC":
        return None, None, True

    def one(part: str) -> str:
        mo, da, yr = (p.strip() for p in part.split("/"))
        y = int(yr)
        y = 2000 + y if y < 100 else y
        return date(y, int(mo), int(da)).isoformat()

    try:
        parts = re.split(r"\s+to\s+", s)
        start = one(parts[0])
        end = one(parts[1]) if len(parts) > 1 else start
        return start, end, False
    except (ValueError, IndexError):
        log.warning("Could not parse event dates: %r", s)
        return None, None, False


def parse_trade_shows(wb) -> list[dict]:
    ws = wb["Events"]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        show = row[1]
        if not show or not str(show).strip():
            continue
        name = _clean(show)
        # Skip the meta/reminder row at the bottom of the tab.
        if name.lower().startswith("post all"):
            continue

        notes = None
        hidden = False
        cadence = re.search(r"\((even|odd) years\)", name, re.I)
        if cadence:
            notes = f"{cadence.group(0).strip('()').capitalize()} only."
            name = _clean(re.sub(r"\((even|odd) years\)", "", name, flags=re.I))

        start, end, needs_confirm = _parse_event_dates(row[2])
        # A show with no date and not flagged TBC doesn't occur in 2027 (e.g. an
        # even-years show in an odd year) — hide it from the scheduler.
        if start is None and not needs_confirm:
            hidden = True
            notes = (notes + " " if notes else "") + "No 2027 occurrence; hidden from scheduler."

        location = _clean(row[3]) if row[3] else None
        if location and location.upper() == "TBC":
            location = None

        out.append(
            {
                "name": name,
                "month": _clean(row[0]) if row[0] else None,
                "start_date": start,
                "end_date": end,
                "location": location,
                "booth": _clean(row[4]) if row[4] else None,
                "link": _clean(row[5]) if row[5] else None,
                "hidden": hidden,
                "needs_date_confirmation": needs_confirm,
                "notes": notes,
            }
        )
    return out


FOOD_KW = ("hot sauce", "poultry", "beef", "pork", "seafood", "fish", "logistics", "meat")
CULTURAL_KW = (
    "lunar new year",
    "ramadan",
    "easter",
    "thanksgiving",
    "happy holidays",
    "christmas",
    "hanukkah",
    "diwali",
    "women's day",
)


def _classify(name: str, is_founding: bool) -> str:
    if is_founding:
        return "globex_founding"
    n = name.lower()
    if any(k in n for k in FOOD_KW):
        return "food_industry"
    if any(k in n for k in CULTURAL_KW):
        return "cultural"
    return "general"


def parse_holidays(wb) -> tuple[list[dict], list[str]]:
    """Holidays tab. Returns (holidays, skipped_names).

    Handles: fill-down month column, 'Entire Month' observances, Easter dedup
    (keep the April-5 row), Globex Founding Day date fix, and dateless skips.
    """
    ws = wb["Holidays"]
    current_month: str | None = None
    holidays: list[dict] = []
    skipped: list[str] = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and str(row[0]).strip() and not str(row[0]).startswith("Holidays &"):
            current_month = _clean(row[0])

        raw_name = row[2]
        if not raw_name or not str(raw_name).strip():
            continue
        name = _clean(raw_name)
        if name.lower() == "holiday":  # header row
            continue

        # Name cleanup + flags.
        date_varies = "(changes)" in name.lower()
        name = _clean(re.sub(r"\(changes\)", "", name, flags=re.I)).rstrip("?").strip()

        c26, c27 = row[3], row[4]
        is_month_long = any(isinstance(c, str) and "entire month" in c.lower() for c in (c26, c27))
        date_2026 = _iso(c26)
        date_2027 = _iso(c27)

        is_founding = "globex founding" in name.lower() or (
            current_month == "November" and date_2026 == "2026-11-05"
        )
        if is_founding:
            name = "Globex Founding Day"
            date_2026, date_2027 = (
                "2026-11-05",
                "2027-11-05",
            )  # fix sheet typo (2027 cell read 2026)

        # Skip dateless, non-month-long rows (e.g. National Fish Day — no date set).
        if not is_month_long and date_2026 is None and date_2027 is None:
            skipped.append(name)
            continue

        descriptions = []
        if date_varies:
            descriptions.append("Date varies year to year.")
        if is_founding:
            descriptions.append("Globex was founded November 5, 1993.")
        if is_month_long:
            descriptions.append("Month-long observance.")

        holidays.append(
            {
                "name": name,
                "month": current_month,
                "date_2026": None if is_month_long else date_2026,
                "date_2027": None if is_month_long else date_2027,
                "is_month_long": is_month_long,
                "category": _classify(name, is_founding),
                "description": " ".join(descriptions) or None,
                "recurring": True,
            }
        )

    # Easter dedup: the sheet lists two "Easter Sunday" rows; keep the one whose
    # 2026 date falls in April (the correct one), drop the March duplicate.
    easters = [h for h in holidays if h["name"].lower().startswith("easter")]
    if len(easters) > 1:
        keep = next(
            (h for h in easters if (h["date_2026"] or "").startswith("2026-04")), easters[0]
        )
        for h in easters:
            if h is not keep:
                holidays.remove(h)
                skipped.append(f"{h['name']} (duplicate {h['date_2026']})")

    return holidays, skipped


# --------------------------------------------------------------------------- #
# Staging
# --------------------------------------------------------------------------- #
def stage_files() -> dict:
    """Copy sample photos + contract; write a chat-reference stub."""
    report = {"photos": 0, "contract": False, "chat_stub": False}

    extract = SOURCE_DIR / "whatsapp_extract"
    if extract.exists():
        for jpg in extract.glob("*PHOTO*.jpg"):
            shutil.copy2(jpg, PHOTO_FIXTURE_DIR / jpg.name)
            report["photos"] += 1

    for pdf in SOURCE_DIR.glob("*Agreement*Final.pdf"):
        shutil.copy2(pdf, DOCS_DIR / "contract.pdf")
        report["contract"] = True
        break

    # Curated Karen-voice excerpts are produced in Phase 2 (prompt tuning); the
    # raw chat stays out of the repo (contains personal numbers). Stub points to it.
    excerpts = DOCS_REFERENCE_DIR / "client_chat_excerpts.md"
    excerpts.write_text(
        "# Client Chat Excerpts (Karen-voice reference)\n\n"
        "> Curated in **Phase 2** for brand-voice prompt tuning. The raw WhatsApp\n"
        "> export is intentionally kept out of the repo (contains personal phone\n"
        f"> numbers). Source: `{extract / '_chat.txt'}`\n\n"
        "_Pending: representative Karen phrasings (approvals, edit requests, tone)._\n",
        encoding="utf-8",
    )
    report["chat_stub"] = True
    return report


def write_missing_assets(employees, shows, holidays, skipped) -> None:
    tbc = [s["name"] for s in shows if s["needs_date_confirmation"]]
    content = f"""# Missing Assets & Outstanding Asks

_Auto-generated by `scripts/import_assets.py` (Phase 0). Send Karen ONE
consolidated request — do not dripfeed._

## Imported OK
- **{len(employees)} employees** (hire dates only; no DOB/age/passwords persisted).
- **{len(shows)} trade shows** ({len(tbc)} with TBC dates).
- **{len(holidays)} holidays** ({sum(h["is_month_long"] for h in holidays)} month-long observances).
- **3 logos** converted to transparent PNG (+ reversed white variants).

## BLOCKING — re-extracted asset ZIP (0-byte in current download)
These block Phase 3 templates and the Phase 6 packaging rotation:
- [ ] `Globex_Animals/` — Cow, Chicken, Pig, Fish illustrations (product spotlights, National Beef/Poultry/Pork/Seafood Month posts)
- [ ] `Globex Packaging/` — all 5 colourways (BLACK, BLUE 288C, RED 1795C, GREEN 349C, + 5th) — the 20 rotating packaging posts
- [ ] `Globex_30 Years/` — 30-year anniversary mark (Nov 5 founding posts, milestone badges)
- [ ] `Grains/` — grain & veggie illustrations (lower priority)

## CONTENT
- [ ] **The 20 rotating brand/packaging posts** — images + per-slot copy/angle, OR free rein for Claude to draft for one-time approval.

## DATA GAPS (Karen fills over time; system degrades gracefully)
- [ ] **TBC trade-show dates** — {len(tbc)} shows: {", ".join(tbc) if tbc else "none"}.
- [ ] **Booth numbers** — all blank in the Events tab.
- [ ] **Dropped (dateless / duplicate) holiday rows** — {", ".join(skipped) if skipped else "none"}. Confirm National Fish Day date or leave dropped.

## LAUNCH (Phase 8)
- [ ] Production Twilio WhatsApp Business sender (Meta approval, 2–7 days) — dedicated number, not Karen's personal line.
- [ ] Karen's number added to `AUTHORIZED_NUMBERS` at handover.
- [ ] Railway project + env vars.
"""
    (DOCS_DIR / "missing_assets.md").write_text(content, encoding="utf-8")


def dump_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    import openpyxl

    for d in (DATA_DIR, LOGO_DIR, PHOTO_FIXTURE_DIR, DOCS_DIR, DOCS_REFERENCE_DIR):
        d.mkdir(parents=True, exist_ok=True)

    if not SOURCE_DIR.exists():
        sys.exit(f"FATAL: source asset dir not found: {SOURCE_DIR}")
    if not EXCEL_PATH.exists():
        sys.exit(f"FATAL: Excel not found: {EXCEL_PATH}")

    gs, magick = setup_tools()

    # --- Logos --------------------------------------------------------------
    log.info("Converting logos (Ghostscript pngalpha -> %dpx PNG)...", TARGET_WIDTH)
    logo_results = []
    for filename, stem in LOGO_MAP.items():
        src = SOURCE_DIR / filename
        if not src.exists():
            log.warning("  MISSING logo source: %s", filename)
            continue
        res = convert_logo(gs, magick, src, stem)
        logo_results.append(res)
        log.info("  %-22s @ %4ddpi -> %s [%s]", stem, res["dpi"], f"{stem}.png", res["full"])

    # --- Excel --------------------------------------------------------------
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    employees = parse_employees(wb)
    shows = parse_trade_shows(wb)
    holidays, skipped = parse_holidays(wb)

    dump_json(DATA_DIR / "employees.json", employees)
    dump_json(DATA_DIR / "trade_shows.json", shows)
    dump_json(DATA_DIR / "holidays.json", holidays)

    # --- Staging + gap report ----------------------------------------------
    staged = stage_files()
    write_missing_assets(employees, shows, holidays, skipped)

    # --- Inventory report ---------------------------------------------------
    tbc = sum(s["needs_date_confirmation"] for s in shows)
    month_long = sum(h["is_month_long"] for h in holidays)
    qualifiers = [
        e["name"] for e in employees if e["hire_date"] and (2026 - int(e["hire_date"][:4])) >= 20
    ]
    log.info("\n" + "=" * 60)
    log.info("PHASE 0 INVENTORY")
    log.info("=" * 60)
    log.info("Logos converted : %d (+ white variants)", len(logo_results))
    log.info("Employees       : %d (hire dates only; no DOB/age/PW)", len(employees))
    log.info("  20+yr milestone: %d -> %s", len(qualifiers), ", ".join(qualifiers))
    log.info("Trade shows     : %d (%d TBC)", len(shows), tbc)
    log.info("Holidays        : %d (%d month-long)", len(holidays), month_long)
    log.info("  dropped rows  : %s", ", ".join(skipped) if skipped else "none")
    log.info(
        "Photos staged   : %d  | contract: %s | chat stub: %s",
        staged["photos"],
        staged["contract"],
        staged["chat_stub"],
    )
    log.info("Outputs -> app/data/*.json, app/templates/assets/logos/*.png, docs/missing_assets.md")


if __name__ == "__main__":
    main()
